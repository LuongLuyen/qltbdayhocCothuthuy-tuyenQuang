from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import logout
from .models import User, Device,BorrowReturn
from .forms import UserForm,DeviceForm,mtForm
from datetime import datetime,timedelta
from django.http import HttpResponse
from openpyxl import Workbook
import pytz

def thongBao(request):
    id = request.session.get('id') 
    thongbao = BorrowReturn.objects.select_related('deviceId','userId').filter(userId=id).order_by('-id')[0:5]
    listT = []
    for x in thongbao:
        if "-T" in x.giaovien:
            listT.append(x)
    return listT

def xuatLichSuMuonTra(list): 
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["ID","Tên thiết bị","Ngày mượn", "Giáo viên mượn cho lớp (-T là Đã trả  )", "Tiết","code","Đơn vị","Giá","Số lượng(mỗi lần)","Ngày nhập","Xuất sứ","Tên người dùng","role","Hạn sử dụng","Trạng thái"] 
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)
    for row_num, device in enumerate(list, start=2):
        worksheet.cell(row=row_num, column=1, value=device.id)
        worksheet.cell(row=row_num, column=2, value=device.deviceId.name)
        worksheet.cell(row=row_num, column=3, value=device.muon)
        worksheet.cell(row=row_num, column=4, value=device.giaovien +" mượn cho: " +device.lop)
        worksheet.cell(row=row_num, column=5, value=device.tiet)
        worksheet.cell(row=row_num, column=6, value=device.deviceId.code)
        worksheet.cell(row=row_num, column=7, value=device.deviceId.unit)
        worksheet.cell(row=row_num, column=8, value=device.deviceId.price)
        worksheet.cell(row=row_num, column=9, value=1)
        worksheet.cell(row=row_num, column=10, value=device.deviceId.date)
        worksheet.cell(row=row_num, column=11, value=device.deviceId.location)
        worksheet.cell(row=row_num, column=12, value=device.userId.name)
        worksheet.cell(row=row_num, column=13, value=device.userId.role)
        worksheet.cell(row=row_num, column=14, value=device.deviceId.hansudung)
        worksheet.cell(row=row_num, column=15, value=device.deviceId.status)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=device_data.xlsx'
    workbook.save(response)
    return response

def xuatHoaChatHetHan(list): 
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["ID","Tên thiết bị", "Môn","code","Đơn vị","Giá","Số lượng","Ngày nhập","Xuất sứ","Hạn sử dụng"] 
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)
    for row_num, device in enumerate(list, start=2):
        worksheet.cell(row=row_num, column=1, value=device.id)
        worksheet.cell(row=row_num, column=2, value=device.name)
        worksheet.cell(row=row_num, column=3, value=device.mon)
        worksheet.cell(row=row_num, column=4, value=device.code)
        worksheet.cell(row=row_num, column=5, value=device.unit)
        worksheet.cell(row=row_num, column=6, value=device.price)
        worksheet.cell(row=row_num, column=7, value=device.quantity)
        worksheet.cell(row=row_num, column=8, value=device.date)
        worksheet.cell(row=row_num, column=9, value=device.location)
        worksheet.cell(row=row_num, column=10, value=device.hansudung)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hsd_data.xlsx'
    workbook.save(response)
    return response

def xuatDanhSachNguoiDung(list):
    workbook = Workbook()
    worksheet = workbook.active
    headers = ["ID","Tên", "Tài khoản","mật khẩu","role"] 
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)
    for row_num, device in enumerate(list, start=2):
        worksheet.cell(row=row_num, column=1, value=device.id)
        worksheet.cell(row=row_num, column=2, value=device.name)
        worksheet.cell(row=row_num, column=3, value=device.userName)
        worksheet.cell(row=row_num, column=4, value=device.password)
        worksheet.cell(row=row_num, column=5, value=device.role)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=hsd_data.xlsx'
    workbook.save(response)
    return response

def timeVietnam(dmy): 
    if dmy == "dmy":
        vietnam = pytz.timezone('Asia/Ho_Chi_Minh')
        time = datetime.now(vietnam)
        return time.strftime('%Y-%m-%d') 
    else:
        vietnam = pytz.timezone('Asia/Ho_Chi_Minh')
        time = datetime.now(vietnam)
        return time.strftime('%Y-%m-%d %H:%M:%S') 

def checkHanSuDung(): 
    list = Device.objects.all()
    device = [] 
    for x in list:
        if x.hansudung != "#":
            device.append(x)
    for x in device:
        dateNow =str(timeVietnam("dmy"))
        hsd= x.hansudung
        if(hsd < dateNow or hsd == dateNow):
            x.status = 'Hết hạn'
            x.save() 

def gioBatDauTiet(tietBD):
    input_time = datetime.strptime(tietBD, "%H:%M:%S")
    result_time = input_time - timedelta(minutes=45)
    result = result_time.strftime("%H:%M:%S")
    return result
def checkGioMuon(): 
    borrowReturn = BorrowReturn.objects.all() 
    listls=[] 
    for x in borrowReturn:
        if "-" in x.giaovien:
            non=0
        else:
            listls.append(x)
    for x in listls:
        result_time_string = gioBatDauTiet(x.tiet)
        T= str(x.muon) + " "+ result_time_string 
        dateNow =str(timeVietnam("no"))
        if T== dateNow or dateNow>T or dateNow< x.tiet:
            device = Device.objects.get(id=x.deviceId_id)
            mt= BorrowReturn.objects.get(id=x.id)
            mt.giaovien = mt.giaovien + "-"
            mt.save() 
            # if int(device.quantity) > 0:
            device.quantity=int(device.quantity) -1
            device.save() 
    for x in borrowReturn:
        result_time_string = gioBatDauTiet(x.tiet)
        T= str(x.muon) + " "+ result_time_string 
        dateNow =str(timeVietnam("no"))
        if dateNow < T:
            if "-" in x.giaovien:
                if "T" in x.giaovien:
                    non=0
                else:
                    device = Device.objects.get(id=x.deviceId_id)
                    mt= BorrowReturn.objects.get(id=x.id)
                    mt.giaovien = x.giaovien[:-1]
                    mt.save() 
                    device.quantity=int(device.quantity) +1
                    device.save()
                    print(mt.giaovien)
def checkLichSuMuon(deviceId,tietm,ngaym): 
    dateNow =str(timeVietnam("no"))
    result_time_string = gioBatDauTiet(tietm)
    T= str(ngaym) + " "+ result_time_string  
    if T < dateNow:
        return False
    device =Device.objects.get(id=deviceId)
    mt =BorrowReturn.objects.filter(deviceId=deviceId) 
    cnt = 1
    slkbd = int(device.quantity) 
    for x in mt:
        if "T" in x.giaovien:
            non=0
        else:
            if "-" in x.giaovien:
                slkbd = int(slkbd) + 1
            datedky= ngaym + " " + tietm
            datex = x.muon + " " + x.tiet
            if datedky == datex :
                cnt = int(cnt) + 1
    if cnt <= slkbd:
        return True
    else:
        return False
    
def chuyenGioSangTiet(list): 
    listmoi=[]
    for x in list:
        if x.tiet == "08:00:00":
            x.tiet= "1"
            listmoi.append(x)
        if x.tiet == "09:05:00":
            x.tiet= "2"
            listmoi.append(x)
        if x.tiet == "09:55:00":
            x.tiet= "3"
            listmoi.append(x)
        if x.tiet == "10:45:00":
            x.tiet= "4"
            listmoi.append(x)
        if x.tiet == "11:35:00":
            x.tiet= "5"
            listmoi.append(x)
    return listmoi
def checkLogin(request):
    id = request.session.get('id') 
    if id != None:
        return True
    else:
        return False
def checkQuyen(request):
    role = request.session.get('role') 
    if role == "ADMIN": 
        return True
    else:
        return False
def loginController(request): 
    rl = checkQuyen(request)
    name = request.session.get('name') 
    id = request.session.get('id') 
    if request.method == 'POST': 
        userName = request.POST.get('userName') 
        password = request.POST.get('password')
        out = request.POST.get('out') 
        xtk = request.POST.get('xtk')
        if(xtk!=None): 
            device = get_object_or_404(User, pk=xtk)
            device.delete()
            return redirect('/')
        l = Device.objects.all()
        device=[]
        for x in l:
            if x.status != "Hết hạn":
                device.append(x)
        if(out=="out"): 
            if 'id' in request.session:
                del request.session['id']
                del request.session['name']
                request.session.pop('role', None)
            logout(request)
            return render(request, 'pages/login.html')
        try:
            user = User.objects.get(userName=userName,password=password) 
            request.session['id'] = str(user.id)
            request.session['role'] = str(user.role)
            request.session['name'] = str(user.name)
            request.session['userName'] = str(user.userName)
            checkHanSuDung() 
            checkGioMuon() 
        except:
            return render(request, 'pages/login.html') 
        listDevice =[] 
        listDevice0 =[] 
        for x in device:
            if x.unit !="phòng" and int(x.quantity) > 0:
                listDevice.append(x)
            if x.unit !="phòng" and int(x.quantity) <= 0:
                listDevice0.append(x)
        id = request.session.get('id') 
        thongbao = BorrowReturn.objects.select_related('deviceId','userId').filter(userId=id).order_by('-id')[0:5] 
        listT = [] 
        for x in thongbao:
            if "-T" in x.giaovien:
                listT.append(x)
        rl = checkQuyen(request)
        name = request.session.get('name') 
        id = request.session.get('id') 
        return render(request, 'pages/home.html',{'device':listDevice,"role":rl,"name":name,"id":id, "device0":listDevice0,"thongbao":listT,"id":id})
    return render(request, 'pages/login.html')

def registerController(request): 
    if request.method == 'POST':
        bq = request.POST.get('bq')
        l = request.POST.get('l')
        form = UserForm(request.POST)
        if l == form.data['password']:
            try:
                user = User.objects.get(userName = form.data['userName'])
                return redirect('/register')     
            except:
                if form.is_valid() and bq !=None:
                    form.save()
                    return redirect('/') 
    else:
        form = UserForm()
    return render(request, 'pages/register.html', {'form': form})
def homeController(request):
    id = request.session.get('id') 
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        name = request.session.get('name') 
        userName = request.session.get('userName') 
        rl = checkQuyen(request)
        if request.method == 'POST':
            deviceId = request.POST.get('deviceId')
            mon = request.POST.get('mon')
            search = request.POST.get('search')
            if search != None: 
                search = search.upper()
                l = Device.objects.all()
                device=[]
                for x in l:
                    if x.status != "Hết hạn":
                        device.append(x)
                list = []
                for x in device:
                    if search in x.name and x.unit != 'phòng':
                        list.append(x)
                    listT =thongBao(request)
                return render(request, 'pages/home.html',{"device": list,"thongbao":listT,"id":id,"name":name,"role":rl,"id":id})
            if deviceId!=None: 
                device = Device.objects.get(id = deviceId)
                listT = thongBao(request)
                return render(request, 'pages/borrowdevice.html',{"device": device,"thongbao":listT,"name":name,"role":rl,"userName":userName})
            if mon!="" and mon!="none": 
                l = Device.objects.all()
                device=[]
                for x in l:
                    if x.status != "Hết hạn":
                        device.append(x)
                list = []
                for x in device:
                    if mon in x.code and x.unit != 'phòng':
                        list.append(x)
                listT = thongBao(request)
                return render(request, 'pages/home.html',{"device": list,"thongbao":listT,"name":name,"role":rl,"id":id})
        l = Device.objects.all()
        device=[]
        for x in l:
            if x.status != "Hết hạn":
                device.append(x)
        listDevice =[] 
        listDevice0 =[]
        for x in device:
            if x.unit !="phòng" and int(x.quantity) > 0:
                listDevice.append(x)
            if x.unit !="phòng" and int(x.quantity) <= 0:
                listDevice0.append(x)
        listT = thongBao(request)
        return render(request, 'pages/home.html',{"device":listDevice,"role":rl,"name":name,"device0":listDevice0,"thongbao":listT,"id":id})
    else:
        return redirect('/') 
def thongkeController(request):
    id = request.session.get('id') 
    if checkLogin(request): 
        checkHanSuDung()
        checkGioMuon()
        nameUser = request.session.get('name') 
        rl = checkQuyen(request)
        l = BorrowReturn.objects.select_related('deviceId','userId').all()
        device =[]
        for x in l:
            if x.deviceId.status != "Hết hạn":
                device.append(x)
        sum =0
        for x in device:
            sum =sum+1
        if request.method == 'POST':
            mon = request.POST.get('mon')
            nam = request.POST.get('nam')
            e = request.POST.get('e')
            end = request.POST.get('end')
            ehsd = request.POST.get('ehsd')
            name = "" 
            nams = ""
            if nam != None:
                nams = nam[0:4]
                name = nam[5:10]
            ngays = request.POST.get('ngays')
            ngaye = request.POST.get('ngaye')
            giaovien = request.POST.get('giaovien')
            if mon != None: # thống kê theo môn
                list = []
                for x in device:
                    if mon in x.deviceId.code:
                        list.append(x)
                    sum =0
                for x in list:
                    sum =sum+1
                listT =thongBao(request)
                return render(request, 'pages/thongke.html',{"device":chuyenGioSangTiet(list), "sum":sum,"name":nameUser,"role":rl,"thongbao":listT,"id":id})
            if giaovien != None: # thống kê theo giáo viên
                list = []
                for x in device:
                    if giaovien.upper() in x.userId.name.upper():
                        list.append(x)
                    sum =0
                for x in list:
                    sum =sum+1
                listT =thongBao(request)
                return render(request, 'pages/thongke.html',{"device":chuyenGioSangTiet(list), "sum":sum,"name":nameUser,"role":rl,"thongbao":listT,"id":id})
            if nams != "" and name != "": # thống kê theo năm
                r = BorrowReturn.objects.filter(muon__gte=f"{nams}-1-1", muon__lte=f"{name}-12-31")
                sum =0
                for x in r:
                    sum =sum+1
                listT =thongBao(request)
                return render(request, 'pages/thongke.html',{"device":chuyenGioSangTiet(r), "sum":sum,"name":nameUser,"role":rl, "thongbao":listT,"id":id})
            if ngays != None and ngaye != None: # thống kê theo ngày nào đến ngày nào
                r = BorrowReturn.objects.filter(muon__gte=f"{ngays}", muon__lte=f"{ngaye}")
                sum =0
                for x in r:
                    sum =sum+1
                listT =thongBao(request)
                return render(request, 'pages/thongke.html',{"device":chuyenGioSangTiet(r),"sum":sum,"name":nameUser,"role":rl,"thongbao":listT,"id":id})
            if e != None: 
                device = BorrowReturn.objects.select_related('deviceId','userId').all().order_by('id')
                file = xuatLichSuMuonTra(chuyenGioSangTiet(device))
                return file
            if end != None: # tương tu
                user = User.objects.all().order_by('id')
                file = xuatDanhSachNguoiDung(user)
                return file
            if ehsd != None: # tuong tu
                device = Device.objects.all().order_by('id')
                listHsd = []
                for x in device:
                    if x.hansudung != "#" and x.status == "Hết hạn":
                        listHsd.append(x)
                file = xuatHoaChatHetHan(listHsd)
                return file
        listT= thongBao(request)
        return render(request, 'pages/thongke.html',{"device":chuyenGioSangTiet(device), "sum":sum,"name":nameUser ,"role":rl,"thongbao":listT,"id":id})
    else:
        return redirect('/')
def adminController(request):
    id = request.session.get('id') 
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        name = request.session.get('name') 
        rl = checkQuyen(request)
        if request.method == 'POST':
            xoa = request.POST.get('xoa')
            capnhat = request.POST.get('capnhat')
            mon = request.POST.get('mon')
            search = request.POST.get('search')
            if search != None and search != '': 
                search=search.upper()
                device = Device.objects.all()
                list = []
                for x in device:
                    if search in x.name:
                        list.append(x)
                listT =thongBao(request)
                return render(request, 'pages/admin.html',{"device": list,"thongbao":listT,"name":name,"role":rl,"id":id})
            if(xoa!=None): 
                device = get_object_or_404(Device, pk=xoa)
                device.delete()
                return redirect('/admins')
            if capnhat != None: 
                device = Device.objects.get(id =capnhat)
                listT =thongBao(request)
                return render(request, 'pages/add.html',{"device":device, "role":rl,"name":name, "thongbao":listT,"id":id})
            if mon!="" and mon != None: 
                device = Device.objects.all()
                listmon =[]
                if mon == "hsd":
                    for x in device:
                        if "Hết hạn" == x.status:
                            listmon.append(x)
                else:
                    for x in device:
                        if mon in x.code:
                            listmon.append(x)
                listT = thongBao(request)
                return render(request, 'pages/admin.html',{"device": listmon,"thongbao":listT,"name":name,"role":rl,"id":id})
        device = Device.objects.all()
        listT = thongBao(request)
        return render(request, 'pages/admin.html',{"device":device,"role":rl,"name":name, "thongbao":listT,"id":id})
    else:
        return redirect('/')
def addController(request):
    if checkLogin(request):
        if request.method == 'POST':
            form = DeviceForm(request.POST)
            update = request.POST.get('capnhat')
            if(update!=None): 
                device = get_object_or_404(Device, pk=update)
                form = DeviceForm(request.POST, instance=device)
                if form.is_valid():
                    form.save()
                    return redirect('/admins')
            if form.is_valid():
                form.save()
                return redirect('/admins')
        id = request.session.get('id') 
        return render(request, 'pages/add.html',{"id":id})
    else:
        return redirect('/')
def labController(request):
    id = request.session.get('id') 
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        userName = request.session.get('userName') 
        name = request.session.get('name') 
        rl = checkQuyen(request)
        if request.method == 'POST':
            deviceId = request.POST.get('deviceId')
            if deviceId != None: 
                device = Device.objects.get(id=deviceId)
                listT = thongBao(request)
                return render(request,'pages/borrowlab.html',{"device": device,"name":name,"role":rl,"userName":userName})
        device = Device.objects.all()
        listDevice1 =[]
        listDevice2 =[]
        listDevice3 =[]
        listDevice4 =[]
        listDeviceKt =[]
        listDevice0 =[]
        for x in device: 
            if x.unit =="phòng":
                if "T1" in x.code :
                    listDevice1.append(x)
                if "T2" in x.code :
                    listDevice2.append(x)
                if "T3" in x.code :
                    listDevice3.append(x)
                if "T4" in x.code :
                    listDevice4.append(x)
                if "KT" in x.code :
                    listDeviceKt.append(x)
            if x.unit =="phòng" and int(x.quantity) <=0:
                mt = BorrowReturn.objects.filter(deviceId=x.id)
                for x in mt:
                    if "-" in x.giaovien:
                        if "T" in x.giaovien:
                            non=0
                        else:
                            listDevice0.append(x)
        
        listT = thongBao(request)
        return render(request, 'pages/lab.html',{"device0":chuyenGioSangTiet(listDevice0),"device1":listDevice1,"device2":listDevice2,"device3":listDevice3,"device4":listDevice4,"deviceKt":listDeviceKt,"role":rl,"name":name,"tb":True,"thongbao":listT,"id":id})
    else:
        return redirect('/')
def borrowLabController(request):
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        userName = request.session.get('userName') 
        if request.method == 'POST':
            giaovien = request.POST.get('giaovien')
            lop = request.POST.get('lop')
            ngaym = request.POST.get('ngaym')
            ngayt = request.POST.get('ngayt')
            tietm = request.POST.get('tietm')
            deviceId = request.POST.get('deviceId')
            id = request.session.get('id') 
            if checkLichSuMuon(deviceId,tietm,ngaym): 
                if giaovien != "" and lop != "" and ngaym != "" and ngayt !="" and tietm != "" and deviceId != "" :
                    borrowReturn = BorrowReturn(userId_id=int(id),deviceId_id=int(deviceId),muon=ngaym,tra=ngayt,lop=lop, giaovien =giaovien,tiet=tietm)
                    borrowReturn.save()
                    return redirect('/thietbidangduocmuon')
            device = Device.objects.get(id = deviceId)
            listT = thongBao(request)
            return render(request, 'pages/borrowLab.html',{"device": device,"thongbao":listT,"userName":userName}) # mấy cái return này trả về view , data tương ứng để hiển thị
        listT = thongBao(request)
        return render(request, 'pages/borrowLab.html',{"thongbao":listT,"userName":userName})
    else:
        return redirect('/')
def borrowDeviceController(request):
    id = request.session.get('id') 
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        userName = request.session.get('userName') 
        if request.method == 'POST':
            giaovien = request.POST.get('giaovien')
            lop = request.POST.get('lop')
            ngaym = request.POST.get('ngaym')
            ngayt = request.POST.get('ngayt')
            tietm = request.POST.get('tietm')
            deviceId = request.POST.get('deviceId')
            update = request.POST.get('update')
            id = request.session.get('id') 
            print(str(ngaym), str(tietm))

            if update == None: 
                if checkLichSuMuon(deviceId,tietm,ngaym):
                    if giaovien != "" and lop != "" and ngaym != "" and ngayt !="" and tietm != "" and deviceId != "":
                        borrowReturn = BorrowReturn(userId_id=int(id),deviceId_id=int(deviceId),muon=ngaym,tra=ngayt,lop=lop, giaovien =giaovien,tiet=tietm)
                        borrowReturn.save()
                        return redirect('/thietbidangduocmuon')
            else: 
                mt = get_object_or_404(BorrowReturn, pk=update)
                form = mtForm(request.POST, instance=mt)
                if checkLichSuMuon(deviceId,form.data['tiet'],form.data['muon']):
                    form.save()
                    return redirect('/thietbidangduocmuon')
                else:
                    mt = BorrowReturn.objects.filter(id=update)
                    return render(request, 'pages/borrowdevice.html',{"devicemt": chuyenGioSangTiet(mt)[0],"userName":userName})
            device = Device.objects.get(id = deviceId)
            listT = thongBao(request)
            return render(request, 'pages/borrowdevice.html',{"device": device,"thongbao":listT,"userName":userName})
        listT = thongBao(request)
        return render(request, 'pages/borrowDevice.html',{"thongbao":listT,"userName":userName})
    else:
        return redirect('/')
def dangMuonController(request):
    id = request.session.get('id') 
    if checkLogin(request):
        checkHanSuDung()
        checkGioMuon()
        userName = request.session.get('userName') 
        name = request.session.get('name') 
        rl = checkQuyen(request)
        if request.method == 'POST':
            mon = request.POST.get('mon')
            xoa = request.POST.get('xoa')
            idtra = request.POST.get('idtra')
            search = request.POST.get('search')
            idxoalich = request.POST.get('idxoalich')
            xoalichsu = request.POST.get('xoalichsu')
            if(xoalichsu!=None): 
                device = get_object_or_404(BorrowReturn, pk=xoalichsu)
                device.delete()
                return redirect('/thietbidangduocmuon')
            if search != None and search != '': 
                search=search.upper()
                idUser = request.session.get('id') 
                device = BorrowReturn.objects.select_related('deviceId','userId').filter(userId=idUser)
                listm=[]
                listt=[]
                listll=[]
                for x in device:
                    if search in x.deviceId.name:
                        if "-" in x.giaovien:
                            if "T" in x.giaovien:
                                listll.append(x)
                            else:
                                listm.append(x)
                        else:
                            listt.append(x)
                    # if "-T" in x.giaovien:
                    #     if search in x.deviceId.name:
                    #         listt.append(x)  
                    # else:
                    #     if search in x.deviceId.name:
                    #         listm.append(x)
                listT =thongBao(request)
                return render(request, 'pages/thietbidangduocmuon.html',{"device1": listm,"device2":listt,"device3":listll,"role":rl,"name":name,"thongbao":listT,"id":id})
            if xoa != None and idtra != None: 
                device = Device.objects.get(id=xoa)
                device.quantity = int(device.quantity)+1
                device.save()
                borrowReturn = BorrowReturn.objects.get(id=idtra)
                borrowReturn.giaovien = str(borrowReturn.giaovien)+"T"
                borrowReturn.save()
                return redirect('/thietbidangduocmuon')
            if idxoalich != None: 
                mt = BorrowReturn.objects.filter(id=idxoalich)
                return render(request, 'pages/borrowdevice.html',{"devicemt": chuyenGioSangTiet(mt)[0],"userName":userName})
            if mon!="": 
                idUser = request.session.get('id') 
                device = BorrowReturn.objects.select_related('deviceId','userId').filter(userId=idUser)
                listm=[] #admin
                listtruT=[] #admin-T
                listTru=[] #admin-
                for x in device:
                    if "-" in x.giaovien:
                        if mon == x.deviceId.code:
                            if "T" in x.giaovien:
                                listtruT.append(x)
                            else:
                                listTru.append(x)
                    else:
                        if mon == x.deviceId.code:
                            listm.append(x)
                listm1 =chuyenGioSangTiet(listm)
                listTru1 =chuyenGioSangTiet(listTru)
                listtruT1 =chuyenGioSangTiet(listtruT)
                listT =thongBao(request)
                return render(request, 'pages/thietbidangduocmuon.html',{"device1": listTru1,"device2":listm1,"device3":listtruT1,"role":rl,"name":name,"thongbao":listT,"id":id})
        idUser = request.session.get('id') 
        device = BorrowReturn.objects.select_related('deviceId','userId').filter(userId=idUser)
        listm=[] #admin
        listtruT=[] #admin-T
        listTru=[] #admin-
        for x in device:
            if "-" in x.giaovien:
                if "T" in x.giaovien:
                    listtruT.append(x)
                else:
                    listTru.append(x)
            else:
                listm.append(x)
        listm1 =chuyenGioSangTiet(listm)
        listTru1 =chuyenGioSangTiet(listTru)
        listtruT1 =chuyenGioSangTiet(listtruT)
        listT = thongBao(request)
        return render(request, 'pages/thietbidangduocmuon.html',{"device1": listTru1,"device2":listm1,"device3":listtruT1,"role":rl,"name":name,"thongbao":listT,"id":id})
    else:
        return redirect('/')
def baseController(request): 
    name = request.session.get('name')
    rl = checkQuyen(request)
    return render(request, 'pages/thongbao.html',{"role": rl,"name":name})